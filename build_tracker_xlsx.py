#!/usr/bin/env python3
"""Build an IBEX-style .xlsx from opportunities.json (mirrors IBEX layout + scoring)."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

opps = json.load(open("opportunities.json"))
opps.sort(key=lambda x: (0 if x.get("status")=="NEW" else 1, -x.get("overall",0)))

wb = Workbook(); ws = wb.active; ws.title = "Funding Tracker"
HDR="2F5496"; SCOREHDR="7030A0"; WEIGHTBG="D9D9D9"; WHITE="FFFFFF"; F="Arial"
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

cols=[("Status",9),("Title",40),("Application",21),("Funding Opp #",17),("Overall",9),
      ("Award Size",8),("Eligibility",9),("Cost Share",8),("Indirect Rates",9),
      ("Proposal Difficulty",11),("Technology Match",11),
      ("Application Deadline",15),("Total Program Funding",16),("Award Size ($)",13),("# Awards",8),
      ("Cost Share Req?",11),("Eligibility (detail)",34),("Link",34),
      ("Agency / Funder",24),("Source",18),("Award Type",15),("Eligibility Channel",16)]
SCORE_FIRST,SCORE_LAST=6,11
WEIGHTS={6:2,7:1,8:1,9:1,10:2,11:3}

ws.merge_cells(start_row=1,start_column=SCORE_FIRST,end_row=1,end_column=SCORE_LAST)
c=ws.cell(1,SCORE_FIRST,"Scoring (1 = Bad, 5 = Good)")
c.font=Font(name=F,bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=SCOREHDR); c.alignment=Alignment(horizontal="center")
ws.cell(1,SCORE_LAST+1,"< score weighting").font=Font(name=F,italic=True,size=9)
for col,w in WEIGHTS.items():
    cell=ws.cell(2,col,w); cell.font=Font(name=F,bold=True); cell.fill=PatternFill("solid",fgColor=WEIGHTBG)
    cell.alignment=Alignment(horizontal="center"); cell.border=border
ws.cell(2,5,"weights:").font=Font(name=F,italic=True,size=9)
for i,(name,width) in enumerate(cols,start=1):
    cell=ws.cell(3,i,name); cell.font=Font(name=F,bold=True,color=WHITE)
    cell.fill=PatternFill("solid",fgColor=SCOREHDR if SCORE_FIRST<=i<=SCORE_LAST else HDR)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
    ws.column_dimensions[get_column_letter(i)].width=width

r=4
for o in opps:
    cs=o.get("cost_share_required")
    vals=[o.get("status",""),o.get("title",""),o.get("application",""),o.get("opp_id",""),
          None,o.get("s_award"),o.get("s_elig"),o.get("s_cost"),o.get("s_indirect"),
          o.get("s_difficulty"),o.get("tech_match"),(o.get("deadline","") or "")[:10],
          o.get("total_funding"),o.get("per_award"),o.get("num_awards"),
          ("No" if cs is False else "Yes" if cs else ""),
          o.get("eligibility_desc",""),o.get("link",""),o.get("funder",""),
          o.get("source",""),o.get("vehicle",""),o.get("eligibility_channel","")]
    for i,v in enumerate(vals,start=1):
        cell=ws.cell(r,i,v); cell.font=Font(name=F,size=10); cell.border=border
        cell.alignment=Alignment(vertical="center",wrap_text=(i in (2,17,18)),
                                 horizontal="center" if (5<=i<=11 or i in(1,15,16)) else "left")
    oc=ws.cell(r,5); oc.value=round(o.get("overall",0),2)
    oc.number_format="0.00"; oc.font=Font(name=F,bold=True,size=10)
    ws.cell(r,13).number_format='$#,##0;;-'
    ws.cell(r,14).number_format='$#,##0;;-'
    # Mark genuinely-absent data clearly (IBEX-style), red italic, so a blank
    # reads as "not posted yet" rather than an error or oversight.
    NA="Not currently available"
    NAFONT=Font(name=F,size=8,italic=True,color="C00000")
    present={
        12: bool((o.get("deadline","") or "").strip()),
        13: o.get("total_funding") not in (None,"",0),
        14: o.get("per_award") not in (None,"",0),
        15: o.get("num_awards") not in (None,"",0),
        16: cs is not None,
        17: bool((o.get("eligibility_desc","") or "").strip()),
    }
    for col,ok in present.items():
        if not ok:
            nc=ws.cell(r,col,NA); nc.font=NAFONT; nc.border=border
            nc.alignment=Alignment(vertical="center",wrap_text=True,
                horizontal="center" if col in (12,13,14,15,16) else "left")
    if o.get("link"):
        lc=ws.cell(r,18); lc.hyperlink=o["link"]; lc.font=Font(name=F,size=9,color="0563C1",underline="single")
    r+=1
last=r-1
if last>=4:
    sc=ColorScaleRule(start_type="num",start_value=1,start_color="F8696B",mid_type="num",mid_value=3,mid_color="FFEB84",end_type="num",end_value=5,end_color="63BE7B")
    ws.conditional_formatting.add(f"F4:K{last}",sc)
    sc2=ColorScaleRule(start_type="num",start_value=2,start_color="F8696B",mid_type="num",mid_value=3.5,mid_color="FFEB84",end_type="num",end_value=5,end_color="63BE7B")
    ws.conditional_formatting.add(f"E4:E{last}",sc2)
    newfill=PatternFill("solid",fgColor="C6EFCE")
    ws.conditional_formatting.add(f"A4:A{last}",CellIsRule(operator="equal",formula=['"NEW"'],fill=newfill,font=Font(name=F,bold=True,color="006100")))
ws.freeze_panes="B4"; ws.row_dimensions[3].height=30

ws2=wb.create_sheet("Methodology")
notes=[("Aethl Funding Tracker — auto-generated",""),("",""),
("Overall Score","(Award Size x2 + Eligibility x1 + Cost Share x1 + Indirect x1 + Proposal Difficulty x2 + Technology Match x3) / 10. Weights editable in row 2."),
("Status","NEW = appeared since last run; ACTIVE = seen before; dropouts logged in opportunities_removed.json."),
("Enriched (real) fields","Total Program Funding, # Awards, Cost Share Req, Eligibility, and the Award Size / Cost Share / Eligibility scores derived from them (Grants.gov detail endpoint)."),
("Technology Match","Scored from opportunity text; LLM scoring planned to remove false positives."),
("Indirect / Proposal Difficulty","Still default assumptions pending Aethl parameters (NICRA rate, grant-writing capacity)."),
("Sources (live)","Grants.gov, EU Funding & Tenders (Horizon/EIC), TED, SAM.gov (when key set)."),
("To refine","Original IBEX .xlsx for exact formatting; award-size bands; eligibility profile; cost-share capacity; indirect rate.")]
for i,(a,b) in enumerate(notes,start=1):
    ws2.cell(i,1,a).font=Font(name=F,bold=(i==1),size=11 if i==1 else 10)
    ws2.cell(i,2,b).font=Font(name=F,size=10); ws2.cell(i,2).alignment=Alignment(wrap_text=True)
ws2.column_dimensions["A"].width=24; ws2.column_dimensions["B"].width=95
wb.save("Aethl_Funding_Tracker.xlsx")
print(f"Wrote tracker with {last-3} opportunities")

